# 读取一张表的一页
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    file = load_workbook(filename=path)  # 打开文件
    sheets = file.get_sheet_names()  # 获取所有的表格的名称
    sheet = file.get_sheet_by_name(sheets[1])   # 获取表格1的数据
    print(sheet.max_row)   # 打印最大行数
    print(sheet.max_column)  # 打印最大列数
    print(sheet.title)  # 打印表名

    # 遍历行数和列数来获取所有的内容
    for lineNum in range(1, sheet.max_row+1):  # 遍历行
        lineList = []  # 定义一个列表用于存储数据
        for columnNum in range(1, sheet.max_column+1):  # 遍历列
            value = sheet.cell(row=lineNum, column=columnNum).value  # sheet.cell函数表示获得文件指定单元格的数据；row=lineNum,column=columnNum分别表示行和列；.value表示获取里面的内容
            # if value != None:  如果加入这句话并与下面一句联立，value为空就不会被添加
            lineList.append(value)
        print(lineList)

path = r"G:\report.xlsx"
readXlsxFile(path)











